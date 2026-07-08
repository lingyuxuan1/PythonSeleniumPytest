import openpyxl
from pathlib import Path
from typing import Any


BASE_DIR = Path(__file__).resolve().parents[1]
DATA_DIR = BASE_DIR / "data"


class get_data:
    def read_execel_login(self) -> list[tuple[Any, ...]]:
        """读取登录 UI 自动化测试数据。"""
        workbook = openpyxl.load_workbook(DATA_DIR / "登录数据.xlsx")
        worksheet = workbook["Sheet1"]
        data = []
        for row in worksheet.iter_rows(min_row=2, values_only=True):
            data.append(row)
        workbook.close()
        return data

    def read_excel_login_data(self) -> list[dict[Any, Any]]:
        """读取登录接口请求测试数据。"""
        workbook1 = openpyxl.load_workbook(DATA_DIR / "登录接口请求测试数据.xlsx")
        worksheet1 = workbook1["Sheet1"]
        data = []
        keys = [cell.value for cell in worksheet1[2]]
        for row in worksheet1.iter_rows(min_row=3, max_row=4, values_only=True):
            dict_data = dict(zip(keys, row))
            data.append(dict_data)
        workbook1.close()
        return data

    def read_excel_skill_data(self) -> list[dict[Any, Any]]:
        """读取技能模块测试数据。"""
        workbook = openpyxl.load_workbook(DATA_DIR / "技能数据.xlsx")
        worksheet = workbook["Sheet1"]
        data = []
        keys = [cell.value for cell in worksheet[1]]
        for row in worksheet.iter_rows(min_row=2, values_only=True):
            if not any(row):
                continue
            data.append(dict(zip(keys, row)))
        workbook.close()
        return data


if __name__ == "__main__":
    a = get_data()
    print(a.read_excel_login_data())
